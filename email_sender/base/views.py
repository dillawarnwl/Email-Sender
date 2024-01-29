from django.core.mail import EmailMultiAlternatives, BadHeaderError, EmailMessage
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from .forms import GivenEmailForm, ExcelEmailForm, StringEmailForm, EmailQueryForm
import openpyxl
# views.py
# from langchain_google_genai import ChatGoogleGenerativeAI
# import os

# MODEL = os.environ.get("MODEL")
# GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY")

# def generate_email_message(query):
#     try:
#         llm = ChatGoogleGenerativeAI(model="gemini-pro")
#         query_with_email_content = f"act as content writer: {query}"
#         result = llm.invoke(query_with_email_content)
#         return result.text if hasattr(result, 'text') else str(result)
#     except Exception as e:
#         # Log the exception or handle it appropriately
#         return f"An error occurred while generating the email message.{e}"

# def email_query(request):
#     if request.method == 'POST':
#         form = EmailQueryForm(request.POST)
#         if form.is_valid():
#             user_query = form.cleaned_data['query']
#             generated_email_message = generate_email_message(user_query)
#             return render(request, 'home.html', {'form': form, 'generated_email_message': generated_email_message})
#     else:
#         form = EmailQueryForm()

#     return render(request, 'home.html', {'form': form})



def from_given_email(request):
    form = GivenEmailForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        subject = form.cleaned_data['subject']
        recipient = form.cleaned_data['dest_email']
        message = form.cleaned_data['message']

        html_content = render_to_string(
            'content_html.html', {'message': message})

        try:
            msg = EmailMultiAlternatives(
                subject, message, 'alazizdonors@gmail.com', [recipient])
            msg.attach_alternative(html_content, "text/html")
            msg.send()
            messages.success(request, "Email sent successfully!")
            return redirect('from_given_email')
        except BadHeaderError:
            return HttpResponse("Invalid header found")
    else:
        context = {'form': form}
        return render(request, 'given_email_form.html', context)


def read_excel_data(excel_file, headers=('email','Email','Email Address')):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Find the column index based on the header
        header_row = sheet[1]
        email_column_index = None

        for idx, cell in enumerate(header_row, start=1):
            if cell.value in headers:
                email_column_index = idx
                break

        if email_column_index is None:
            raise ValueError(f"Column with header '{headers}' not found in the Excel sheet")

        email_addresses = []

        # Iterate through rows and extract email addresses
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming data starts from the second row
            email = row[email_column_index - 1]  # Adjust index since iter_rows starts from 1
            if email:
                email_addresses.append(email)

        return email_addresses

    except Exception as e:
        # Log or handle the exception as needed
        return f"Error reading Excel data: {e}"



def send_emails(request, email_addresses, subject, message):
    try:
        for email_address in email_addresses:
            html_content = render_to_string(
                'blood_invitation.html', {'message': message})
            email = EmailMessage(
                subject,
                html_content,
                'alazizdonors@gmail.com',
                [email_address],
            )
            email.content_subtype = 'html'
            email.send()
        messages.success(request, 'Emails sent successfully.')
    except Exception as e:
        # Log or handle the exception as needed
        messages.error(request, f'Error sending emails.{e}')


def from_excel(request):
    if request.method == 'POST':
        form = ExcelEmailForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']
            email_addresses = read_excel_data(excel_file)
            print(email_addresses)

            # send_emails(request, email_addresses, subject, message)
            return redirect('from_given_email')

    else:
        form = ExcelEmailForm()

    context = {'form': form}
    return render(request, 'excel_form.html', context)


def from_string(request):
    form = StringEmailForm(request.POST or None)
    if form.is_valid():
        email_string = form.cleaned_data['dest_email']
        subject = form.cleaned_data['subject']
        message = form.cleaned_data['message']

        email_addresses = [email.strip()
                           for email in email_string.split(',') if email.strip()]
        send_emails(request, email_addresses, subject, message)
        return redirect('from_given_email')

    context = {'form': form}
    return render(request, 'string_form.html', context)